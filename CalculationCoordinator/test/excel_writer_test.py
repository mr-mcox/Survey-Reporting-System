import unittest
from SurveyReportingSystem.CalculationCoordinator import CalculationCoordinator
import pandas as pd
import os, sys
from openpyxl import load_workbook
from openpyxl import Workbook
from unittest import mock
from SurveyReportingSystem.ConfigurationReader import ConfigurationReader

class WriteExcelTestCase(unittest.TestCase):

	def setUp(self):
		# assert not os.path.exists('test_file.xlsx')
		wb = Workbook()
		ws = wb.get_active_sheet()
		wb.create_named_range('disp_value_col_head',ws,'$A$1')
		ws.title = "ExistingData"
		wb.save('test_file.xlsx')
		coordinator = CalculationCoordinator.CalculationCoordinator()#responses=pd.DataFrame({'response':[1],'respondent_id':[1]})
		coordinator.compute_historical = True
		self.mapping_values = ['dos','uno','tres']
		self.mapping_integers = ['2','1','3']
		self.mapping_dict = dict(zip(self.mapping_integers,self.mapping_values))
		self.labels_for_cut_dimensions = {
											'gender':{'male':'0','female':'1'},
											'region':{'Atlanta':'2','SoDak':'3'}
										}
		config_reader = ConfigurationReader.ConfigurationReader()
		config_reader.config['excel_template_file'] = 'test_file.xlsx'
		config_reader.config['cut_menu_order'] = ['Region','Subject','Grade']
		self.cuts_for_excel_menu_responses = [['Grade', 'static', 'Grade', 'Region', 'Corps'], ['Region', 'static', 'Region', 'Corps','None'],['Gender', 'static', 'Gender', 'Region', 'Corps']]
		return_for_cuts_for_excel_menu = {None:self.cuts_for_excel_menu_responses,'historical':[['Region', 'static', 'Region', 'Corps','None']],'cuts_2':[['Gender', 'static', 'Gender', 'Region', 'Corps']],'cuts_3':[],'cuts_4':[],'cuts_5':[]}
		config_reader.cuts_for_excel_menu = mock.MagicMock(side_effect= lambda **arg: return_for_cuts_for_excel_menu[arg['menu']])
		coordinator.config = config_reader

		dimension_with_all_together_label = ConfigurationReader.Dimension(title="GenderB")
		dimension_with_all_together_label.all_together_label = "Gender Not Used"
		dimension_with_dynamic_type = ConfigurationReader.Dimension(title="GenderC")
		dimension_with_dynamic_type.dimension_type = "dynamic"
		dimension_with_dynamic_type.dynamic_parent_dimension = "region"

		dimension_with_composites = ConfigurationReader.Dimension(title="Gender_LIB")
		dimension_with_composites.is_composite = True
		dimension_with_composites.composite_dimensions = ['Gender','LIB']

		dimension_with_value_order = ConfigurationReader.Dimension(title="ValueOrderDimension")
		dimension_with_value_order.value_order = ['B','A']

		config_reader.all_dimensions = mock.MagicMock(return_value = [
																		ConfigurationReader.Dimension(title="Gender"),
																		dimension_with_all_together_label,
																		dimension_with_dynamic_type,
																		dimension_with_composites,
																		dimension_with_value_order,
																		])
		coordinator.get_integer_string_mapping = mock.MagicMock(return_value= {'integer_strings':['1','2'],'labels':['male','female']})

		#TODO: CREATE NEW ONE FOR LIB
		generic_gender_return = {'integer_strings':['1','2'],'labels':['male','female']}
		value_order_return = {'integer_strings':['1','2','3'],'labels':['A','B','C']}
		lib_return = {'integer_strings':['4'],'labels':['lib']}
		return_for_get_integer_string_mapping = {'Gender':generic_gender_return,
												'GenderB':generic_gender_return,
												'GenderC':generic_gender_return,
												'question_code':generic_gender_return,
												'result_type':generic_gender_return,
												'LIB': lib_return,
												'ValueOrderDimension':value_order_return,
												'survey_code':generic_gender_return,}
		coordinator.get_integer_string_mapping = mock.MagicMock(side_effect= lambda arg: return_for_get_integer_string_mapping[arg])

		coordinator.dimension_integer_mapping = {'values': self.mapping_values,'integers':self.mapping_integers}
		coordinator.excel_dashboard_file = 'test_file.xlsx'
		coordinator.labels_for_cut_dimensions = self.labels_for_cut_dimensions
		master_aggregation = pd.DataFrame({
				'row_heading':['0','0','1','1'],
				'column_heading': ['2;3','2;4','2;3','2;4'],
				'aggregation_value': [0.5,0.5,0.5,0.5]
			})
		coordinator.setup_flush_file('df_dv.csv')
		coordinator.flush_aggregation_to_file('df_dv.csv',master_aggregation)
		coordinator.setup_flush_file('df_dv_hist.csv')
		coordinator.flush_aggregation_to_file('df_dv_hist.csv',master_aggregation)
		coordinator.compute_cuts_from_config = mock.MagicMock(return_value=master_aggregation)

		master_siginificance = pd.DataFrame({
				'row_heading':['0','0','1','1'],
				'column_heading': ['2;3','2;4','2;3','2;4'],
				'aggregation_value': ['H','','','S']
		})
		coordinator.setup_flush_file('df_sig.csv')
		coordinator.flush_aggregation_to_file('df_sig.csv',master_siginificance)
		coordinator.setup_flush_file('df_sig_hist.csv')
		coordinator.flush_aggregation_to_file('df_sig_hist.csv',master_siginificance)

		#Set up demographics for dynamic columns
		coordinator.demographic_data = pd.DataFrame({'respondent_id':[1,2,3],'region':['Atlanta','Atlanta','SoDak'],'GenderC':['male','female','female']})
		coordinator.responses = pd.DataFrame({'respondent_id':[1,2,3],'response':[1,2,1]})
		coordinator.compute_significance_from_config = mock.MagicMock(return_value=master_siginificance)
		coordinator.export_to_excel()
		self.coordinator = coordinator

	def test_writing_cut_config(self):
		ws = load_workbook(filename = r'test_file.xlsx').get_sheet_by_name(name = 'Lookups')
		for i in range(2):
			row_values = [str(ws.cell(row=i+1,column=j+1).value) for j in range(5)]
			row_matches_expected = False
			for expected_row in self.cuts_for_excel_menu_responses:
				if expected_row == row_values:
					row_matches_expected = True
			print(row_values)
			print(expected_row)
			self.assertTrue(row_matches_expected)

	def test_cut_menu_order(self):
		ws = load_workbook(filename = r'test_file.xlsx').get_sheet_by_name(name = 'Lookups')
		assert ws.cell(row=1,column=6).value == "Region"
		assert ws.cell(row=2,column=6).value == "Grade"
		assert ws.cell(row=3,column=6).value == "Gender"

	def test_other_cuts_menu(self):
		ws = load_workbook(filename = r'test_file.xlsx').get_sheet_by_name(name = 'Lookups')
		assert ws.cell(row=4,column=6).value == "Region"
		assert ws.cell(row=5,column=6).value == "Gender"

	def test_writing_menu_translations(self):
		ws = load_workbook(filename = r'test_file.xlsx').get_sheet_by_name(name = 'Lookups')
		assert ws.cell(row=1,column=7).value == "Gender"
		assert ws.cell(row=2,column=7).value == "male"
		assert ws.cell(row=3,column=7).value == "female"
		assert ws.cell(row=2,column=8).value == str(1)
		assert ws.cell(row=3,column=8).value == str(2)


	def test_writing_menu_for_no_all_together_label(self):
		ws = load_workbook(filename = r'test_file.xlsx').get_sheet_by_name(name = 'Lookups')
		assert ws.cell(row=1,column=9).value == "GenderB"
		assert ws.cell(row=2,column=9).value == "Gender Not Used"
		assert ws.cell(row=3,column=9).value == "male"
		assert ws.cell(row=4,column=9).value == "female"
		assert ws.cell(row=2,column=10).value == str(0)
		assert ws.cell(row=3,column=10).value == str(1)
		assert ws.cell(row=4,column=10).value == str(2)

	def test_writing_menu_for_dynamic_dimension(self):
		ws = load_workbook(filename = r'test_file.xlsx').get_sheet_by_name(name = 'Lookups')
		assert ws.cell(row=1,column=11).value == "GenderC"
		assert ws.cell(row=2,column=11).value == "female"
		assert ws.cell(row=3,column=11).value == "male"
		assert ws.cell(row=4,column=11).value == "female"
		assert ws.cell(row=2,column=12).value == str(2)
		assert ws.cell(row=3,column=12).value == str(1)
		assert ws.cell(row=4,column=12).value == str(2)
		assert ws.cell(row=2,column=13).value == "Atlanta"
		assert ws.cell(row=3,column=13).value == "Atlanta"
		assert ws.cell(row=4,column=13).value == "SoDak"

	def test_write_master_as_excel(self):
		ws = load_workbook(filename = r'test_file.xlsx').get_sheet_by_name(name = 'DisplayValues')
		column_headings = [str(ws.cell(row=1,column=i+2).value) for i in range(ws.get_highest_column()-1)]
		self.assertEqual(set(column_headings), {'2;3','2;4'}) 
		row_headings = [str(ws.cell(row=i+2,column=1).value) for i in range(ws.get_highest_row()-1)]
		self.assertEqual(set(row_headings), {'0','1'})

	def test_write_significance_values(self):
		ws = load_workbook(filename = r'test_file.xlsx').get_sheet_by_name(name = 'SignificanceValues')
		column_headings = [str(ws.cell(row=1,column=i+2).value) for i in range(ws.get_highest_column()-1)]
		self.assertEqual(set(column_headings), {'2;3','2;4'}) 
		row_headings = [str(ws.cell(row=i+2,column=1).value) for i in range(ws.get_highest_row()-1)]
		self.assertEqual(set(row_headings), {'0','1'})

	def test_write_hist_master_as_excel(self):
		ws = load_workbook(filename = r'test_file.xlsx').get_sheet_by_name(name = 'HistDisplayValues')
		column_headings = [str(ws.cell(row=1,column=i+2).value) for i in range(ws.get_highest_column()-1)]
		self.assertEqual(set(column_headings), {'2;3','2;4'}) 
		row_headings = [str(ws.cell(row=i+2,column=1).value) for i in range(ws.get_highest_row()-1)]
		self.assertEqual(set(row_headings), {'0','1'})

	def test_write_hist_significance_values(self):
		ws = load_workbook(filename = r'test_file.xlsx').get_sheet_by_name(name = 'HistSignificanceValues')
		column_headings = [str(ws.cell(row=1,column=i+2).value) for i in range(ws.get_highest_column()-1)]
		self.assertEqual(set(column_headings), {'2;3','2;4'}) 
		row_headings = [str(ws.cell(row=i+2,column=1).value) for i in range(ws.get_highest_row()-1)]
		self.assertEqual(set(row_headings), {'0','1'})

	def test_named_ranges_on_display_values(self):
		wb = load_workbook(filename = r'test_file.xlsx')
		range_names = [r.name for r in wb.get_named_ranges()]
		self.assertTrue( 'disp_value_row_head' in range_names)
		self.assertTrue( 'disp_value_col_head' in range_names)
		self.assertTrue( 'disp_value_values' in range_names)
		self.assertEqual( wb.get_named_range('disp_value_row_head').destinations[0][1],'$A$2:$A$3')
		self.assertEqual( wb.get_named_range('disp_value_col_head').destinations[0][1],'$B$1:$C$1')
		self.assertEqual( wb.get_named_range('disp_value_values').destinations[0][1],'$B$2:$C$3')

	def test_named_ranges_on_significance_values(self):
		wb = load_workbook(filename = r'test_file.xlsx')
		range_names = [r.name for r in wb.get_named_ranges()]
		self.assertTrue( 'sig_value_row_head' in range_names)
		self.assertTrue( 'sig_value_col_head' in range_names)
		self.assertTrue( 'sig_value_values' in range_names)
		self.assertEqual( wb.get_named_range('sig_value_row_head').destinations[0][1],'$A$2:$A$3')
		self.assertEqual( wb.get_named_range('sig_value_col_head').destinations[0][1],'$B$1:$C$1')
		self.assertEqual( wb.get_named_range('sig_value_values').destinations[0][1],'$B$2:$C$3')

	def test_named_ranges_on_hist_display_values(self):
		wb = load_workbook(filename = r'test_file.xlsx')
		range_names = [r.name for r in wb.get_named_ranges()]
		self.assertTrue( 'hist_disp_value_row_head' in range_names)
		self.assertTrue( 'hist_disp_value_col_head' in range_names)
		self.assertTrue( 'hist_disp_value_values' in range_names)
		self.assertEqual( wb.get_named_range('hist_disp_value_row_head').destinations[0][1],'$A$2:$A$3')
		self.assertEqual( wb.get_named_range('hist_disp_value_col_head').destinations[0][1],'$B$1:$C$1')
		self.assertEqual( wb.get_named_range('hist_disp_value_values').destinations[0][1],'$B$2:$C$3')

	def test_named_ranges_on_hist_significance_values(self):
		wb = load_workbook(filename = r'test_file.xlsx')
		range_names = [r.name for r in wb.get_named_ranges()]
		self.assertTrue( 'hist_sig_value_row_head' in range_names)
		self.assertTrue( 'hist_sig_value_col_head' in range_names)
		self.assertTrue( 'hist_sig_value_values' in range_names)
		self.assertEqual( wb.get_named_range('hist_sig_value_row_head').destinations[0][1],'$A$2:$A$3')
		self.assertEqual( wb.get_named_range('hist_sig_value_col_head').destinations[0][1],'$B$1:$C$1')
		self.assertEqual( wb.get_named_range('hist_sig_value_values').destinations[0][1],'$B$2:$C$3')

	def test_named_ranges_on_lookup_tab(self):
		wb = load_workbook(filename = r'test_file.xlsx')
		range_names = [r.name for r in wb.get_named_ranges()]
		self.assertTrue( 'cuts' in range_names)
		self.assertTrue( 'cuts_config' in range_names)
		self.assertTrue( 'default_menu' in range_names)
		self.assertTrue( 'cuts_head' in range_names)
		self.assertEqual( wb.get_named_range('cuts').destinations[0][1],'$F$1:$F$3')
		self.assertEqual( wb.get_named_range('cuts_historical').destinations[0][1],'$F$4:$F$4')
		self.assertEqual( wb.get_named_range('cuts_2').destinations[0][1],'$F$5:$F$5')
		self.assertEqual( wb.get_named_range('cuts_config').destinations[0][1],'$A$1:$E$3')
		self.assertEqual( wb.get_named_range('default_menu').destinations[0][1],'$F$2:$F$1001')
		self.assertEqual( wb.get_named_range('default_mapping').destinations[0][1],'$F$2:$G$1001')
		self.assertEqual( wb.get_named_range('default_menu_start').destinations[0][1],'$F$2')
		self.assertEqual( wb.get_named_range('cuts_head').destinations[0][1],'$G$1:$W$1')

	def test_composite_dimensions(self):
		ws = load_workbook(filename = r'test_file.xlsx').get_sheet_by_name(name = 'Lookups')
		assert ws.cell(row=1,column=14).value == "Gender_LIB"
		assert ws.cell(row=2,column=14).value == "male"
		assert ws.cell(row=3,column=14).value == "female"
		assert ws.cell(row=4,column=14).value == "lib"

	def test_value_order_dimension(self):
		ws = load_workbook(filename = r'test_file.xlsx').get_sheet_by_name(name = 'Lookups')
		assert ws.cell(row=1,column=16).value == "ValueOrderDimension"
		assert ws.cell(row=2,column=16).value == "B"
		assert ws.cell(row=3,column=16).value == "A"
		
	def test_zero_string_label(self):
		wb = load_workbook(filename = r'test_file.xlsx')
		ws = load_workbook(filename = r'test_file.xlsx').get_sheet_by_name(name = 'Lookups')
		range_names = [r.name for r in wb.get_named_ranges()]
		self.assertTrue( 'zero_string' in range_names)
		self.assertEqual( ws.cell(wb.get_named_range('zero_string').destinations[0][1] ).value, str(0) )

	def test_question_code_mapping_provided(self):
		ws = load_workbook(filename = r'test_file.xlsx').get_sheet_by_name(name = 'Lookups')
		top_row_headings = [str(ws.cell(row=1,column=i+1).value) for i in range(ws.get_highest_column())]
		self.assertTrue('question_code' in top_row_headings)

	def test_result_type_mapping_provided(self):
		ws = load_workbook(filename = r'test_file.xlsx').get_sheet_by_name(name = 'Lookups')
		top_row_headings = [str(ws.cell(row=1,column=i+1).value) for i in range(ws.get_highest_column())]
		self.assertTrue('result_type' in top_row_headings)

	def test_template_sheets_untouched(self):
		wb = load_workbook(filename = r'test_file.xlsx')
		self.assertTrue('ExistingData' in wb.get_sheet_names())

	def test_ensure_combination_for_every_set_of_demographics_is_true(self):
		self.assertTrue(self.coordinator.ensure_combination_for_every_set_of_demographics)

	def tearDown(self):
		try:
			pass
			# os.remove('test_file.xlsx')
		except:
			pass


if __name__ == '__main__':
    unittest.main()