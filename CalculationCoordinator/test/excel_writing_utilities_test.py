import unittest
from SurveyReportingSystem.CalculationCoordinator import CalculationCoordinator
from openpyxl import load_workbook, Workbook
import os

class ExcelWritingUtilityTestCase(unittest.TestCase):
	
	def setUp(self):
		wb1 = Workbook()
		ws1 = wb1.get_active_sheet()
		ws1.cell(row=0,column=0).value = "WB1Value"
		ws1.title = "WB1Sheet"
		wb1.save('wb1.xlsx')

	def test_copy_sheet_between_wb(self):
		wb2 = Workbook()
		ws2 = wb2.get_active_sheet()
		ws2.cell(row=0,column=0).value = "WB2Value"
		ws2.title = "WB2Sheet"
		wb2.save('wb2.xlsx')

		cc = CalculationCoordinator()
		cc.copy_sheet_to_workbook('wb1.xlsx','WB1Sheet','wb2.xlsx')

		orig_sheet = load_workbook(filename = r'wb2.xlsx').get_sheet_by_name(name = 'WB2Sheet')
		test_sheet = load_workbook(filename = r'wb2.xlsx').get_sheet_by_name(name = 'WB1Sheet')
		self.assertEqual(test_sheet.cell(row=0,column=0).value,'WB1Value')
		self.assertEqual(orig_sheet.cell(row=0,column=0).value,'WB2Value')

	def test_replace_sheet_between_wb(self):
		wb2 = Workbook()
		ws2 = wb2.get_active_sheet()
		ws2.cell(row=0,column=0).value = "WB2Value"
		ws2.title = "WB1Sheet"
		wb2.save('wb2.xlsx')

		cc = CalculationCoordinator()
		cc.copy_sheet_to_workbook('wb1.xlsx','WB1Sheet','wb2.xlsx')

		test_sheet = load_workbook(filename = r'wb2.xlsx').get_sheet_by_name(name = 'WB1Sheet')
		self.assertEqual(test_sheet.cell(row=0,column=0).value,'WB1Value')

	def tearDown(self):
		try:
			pass
			os.remove('wb1.xlsx')
			os.remove('wb2.xlsx')
		except:
			pass