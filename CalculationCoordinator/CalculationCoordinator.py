import pandas as pd
from SurveyReportingSystem.NumericOutputCalculator import NumericOutputCalculator
from SurveyReportingSystem.ConfigurationReader import ConfigurationReader
import math
from openpyxl import load_workbook, cell, Workbook
import copy
import logging
import os
import gc
import re
import csv
import numpy as np
import pdb

class CalculationCoordinator(object):
	"""docstring for CalculationCoordinator"""
	def __init__(self, **kwargs):
		self.responses = kwargs.pop('responses',None)
		self.demographic_data = kwargs.pop('demographic_data',None)
		if self.demographic_data is not None:
			self.demographic_data = self.demographic_data.set_index('respondent_id').applymap(str).fillna("Missing").reset_index()
		self.hist_responses = kwargs.pop('hist_responses',None)
		self.hist_demographic_data = kwargs.pop('hist_demographic_data',None)
		if self.hist_demographic_data is not None:
			self.hist_demographic_data = self.hist_demographic_data.set_index('respondent_id').applymap(str).fillna("Missing").reset_index()
		self.config = kwargs.pop('config',None)
		self.computations_generated = dict()
		self.result_types = kwargs.pop('result_types',['net'])
		self.labels_for_cut_dimensions = dict()
		self.integers_for_cut_dimensions = dict()
		self.max_integer_used_for_integer_string = 1
		self.zero_integer_string = '0'
		self.ensure_combination_for_every_set_of_demographics = False
		self.integer_string_length = math.floor(math.log(self.max_integer_used_for_integer_string,10) + 1 )
		self.format_string = "{0:0" + str(self.integer_string_length) + "d}"
		self.zero_integer_string = self.format_string.format(0)

		self.compute_historical = False
		if self.hist_responses is not None and self.hist_demographic_data is not None:
			self.compute_historical = True

	def get_aggregation(self,**kwargs):
		cuts = kwargs.pop('cuts',None)
		
		if type(cuts) != list:
			cuts = [cuts]
		aggregation_key = tuple(cuts)

		assert aggregation_key in self.computations_generated
		return self.computations_generated[aggregation_key]

	# @profile
	def compute_aggregation(self,**kwargs):
		responses = kwargs.pop('responses',self.responses)
		demographic_data = kwargs.pop('demographic_data',self.demographic_data)
		calculator = kwargs.pop('calculator',NumericOutputCalculator.NumericOutputCalculator(responses=responses,demographic_data=demographic_data,responses_transformed=True))
		assert type(self.result_types) == list
		orig_cuts = kwargs.pop('cut_demographic',None)
		if type(orig_cuts) != list:
			orig_cuts = [orig_cuts]
		cuts = [cut for cut in orig_cuts if cut != None]

		calculations = calculator.compute_aggregation(result_type=self.result_types,cut_demographic=cuts,**kwargs)
		gc.collect()

		if self.ensure_combination_for_every_set_of_demographics and len(cuts) > 0:
			# logging.debug("Cuts passing to modify " + str(cuts))
			calculations = self.modify_for_combinations_of_demographics(df=calculations,cuts=cuts,demographic_data=demographic_data)

		aggregation_key = tuple(cuts)

		gc.collect()
		self.computations_generated[aggregation_key] = calculations
		assert 'level_0' not in calculations.columns
		return calculations

	# @profile
	def compute_significance(self,**kwargs):
		responses = kwargs.pop('responses',self.responses)
		demographic_data = kwargs.pop('demographic_data',self.demographic_data)
		calculator = kwargs.pop('calculator',NumericOutputCalculator.NumericOutputCalculator(responses=responses,demographic_data=demographic_data,responses_transformed=True))
		assert type(self.result_types) == list
		orig_cuts = kwargs.pop('cut_demographic',None)
		if type(orig_cuts) != list:
			orig_cuts = [orig_cuts]
		cuts = [cut for cut in orig_cuts if cut != None]

		calculations = calculator.bootstrap_net_significance(cuts=cuts,**kwargs).reset_index()
		gc.collect()

		if self.ensure_combination_for_every_set_of_demographics and len(cuts) > 0:
			logging.debug("Cuts passing to modify " + str(cuts))
			calculations = self.modify_for_combinations_of_demographics(df=calculations,cuts=cuts,demographic_data=demographic_data)

		aggregation_key = tuple(cuts)

		gc.collect()
		assert 'level_0' not in calculations.columns
		# self.computations_generated[aggregation_key] = calculations
		return calculations

	# @profile
	def modify_for_combinations_of_demographics(self, df, cuts,**kwargs):
		assert 'level_0' not in df.columns
		# logging.debug("Index name is " + str(df.index.name))
		# logging.debug("Sampling of df:\n"+str(df.head()))
		# # df = df.reset_index()
		# logging.debug("Sampling of df after reset:\n"+str(df.head()))
		# assert 'level_0' not in df.columns
		# logging.debug("Starting df is " + str(df.head()))
		# logging.debug("Starting df row is " + str(df.reset_index().ix[0,:]))
		demographic_data = kwargs.pop('demographic_data',self.demographic_data)
		logging.debug('Modifying cuts\n' + str(cuts) + '\ndf is \n' + str(df.head()))
		first_result_type = df['result_type'].iloc[0]
		# logging.debug("First result type is " + df.ix[0,'result_type'])
		first_question_code = df['question_code'].iloc[0]
		if hasattr(self,'default_question'):
			first_question_code = self.default_question
		if hasattr(self,'default_result_type'):
			first_result_type = self.default_result_type

		#Initialize with first list
		assert len(cuts) > 0
		# logging.debug("Cuts are " + str(cuts))
		assert cuts[0] in demographic_data, cuts[0] + " not found in demographic_data"
		master_list_of_demographics = [ [item] for item in demographic_data[cuts[0]].unique().tolist()]

		#Create list of combinations to look for
		for i in range(len(cuts) - 1):
			cur_cut = cuts[i+1]
			assert cur_cut in demographic_data
			cur_cut_values = demographic_data[cur_cut].unique().tolist()
			value_set_for_cut = list()
			for value in cur_cut_values:
				list_to_add_onto = copy.deepcopy(master_list_of_demographics)
				for item in list_to_add_onto:
					item.append(value)
				value_set_for_cut += list_to_add_onto
			master_list_of_demographics = value_set_for_cut

		#Convert list to tuples
		master_list_of_demographics = {tuple(value_set) for value_set in master_list_of_demographics}

		#Remove all rows where dynamic dimension pairs don't exist
		if self.config is not None:
				all_dimensions = {dimension.title: dimension for dimension in self.config.all_dimensions()}
				position_of_cut = {cut: i for i, cut in enumerate(cuts)}
				for cur_cut in position_of_cut.keys():
					if cur_cut in all_dimensions and all_dimensions[cur_cut].dimension_type == 'dynamic' and all_dimensions[cur_cut].dynamic_parent_dimension in cuts:
						parent_dimension = all_dimensions[cur_cut].dynamic_parent_dimension
						position_of_parent = position_of_cut[parent_dimension]
						position_of_child = position_of_cut[cur_cut]
						df_for_dynamic = df.copy().set_index([parent_dimension,cur_cut])
						value_sets_to_remove = set()
						for value_set in master_list_of_demographics:
							dimension_combo_to_check_for = (value_set[position_of_parent],value_set[position_of_child])
							if df_for_dynamic.index.isin(dimension_combo_to_check_for).sum() == 0:
								value_sets_to_remove.add(value_set)
						for value_set in value_sets_to_remove:
							master_list_of_demographics.remove(value_set)

		#Add new rows where necessary
		# df_with_cut_index = df.copy().set_index(cuts)
		all_value_sets = pd.DataFrame(list(master_list_of_demographics))
		if len(master_list_of_demographics) == 0:
			return df
		else:
			all_value_sets.columns = cuts
			# all_value_sets['result_type'] = first_result_type
			# all_value_sets['question_code'] = first_question_code
			all_value_sets = all_value_sets.set_index(cuts)
			df_with_additional_rows = df.set_index(cuts).join(all_value_sets,how="outer")
			df_with_additional_rows.result_type.fillna(first_result_type,inplace=True)
			df_with_additional_rows.question_code.fillna(first_question_code,inplace=True)
			# for value_set in master_list_of_demographics:
			# 	if df_with_cut_index.index.isin([tuple(value_set)]).sum()== 0:
			# 		dict_for_df = dict()
			# 		dict_for_df['result_type'] = [first_result_type]
			# 		# logging.debug("First question code is " + first_question_code )
			# 		dict_for_df['question_code'] = [first_question_code]
			# 		new_row = pd.DataFrame(dict_for_df)
			# 		# logging.debug("New row before adding cuts:\n" + str(new_row))
			# 		for i in range(len(cuts)):
			# 			new_row[cuts[i]] = value_set[i]
			# 		# logging.debug("New row after adding cuts:\n" + str(new_row))

			# 		df = pd.concat([df,new_row])
			return df_with_additional_rows.reset_index()

	def replace_dimensions_with_integers(self, df = None):
		assert 'level_0' not in df.columns

		values_by_column = self.integers_for_cut_dimensions
		#Collect all values by column
		for column in df.columns:
			if column != 'aggregation_value':
				if column not in values_by_column:
					values_by_column[column] = dict()
				for value in df[column].unique():
					if value not in values_by_column[column]:
						values_by_column[column][value] = self.max_integer_used_for_integer_string
						self.max_integer_used_for_integer_string += 1

		self.integer_string_length = math.floor(math.log(self.max_integer_used_for_integer_string,10) + 1 )

		self.format_string = "{0:0" + str(self.integer_string_length) + "d}"
		self.zero_integer_string = self.format_string.format(0)

		integer_strings_by_column = { column_key : { value : self.format_string.format(x) for (value, x) in values_dict.items()} for (column_key,values_dict) in values_by_column.items()}

		#Map values to column
		for column in df.columns:
			if column != 'aggregation_value' and column != 'index':
				value_list = [value for (key, value) in integer_strings_by_column[column].items()]
				key_list = [key for (key, value) in integer_strings_by_column[column].items()]
				value_map = pd.Series(value_list,index=key_list)
				assert value_map.index.is_unique, "Duplicate values for column " + column + " include\n" + str(value_map.groupby(level=0).filter(lambda x: len(x) > 1))
				# logging.debug("Mapping values for column " + column + "\n" + str(value_map))
				# logging.debug("Here's the column\n"+str(df.loc[:,column].head()))
				# logging.debug("Here's the value map\n"+str(value_map))
				# logging.debug("Here's the return\n"+str(df.loc[:,column].map(value_map).head()))
				df.loc[:,column] = df.loc[:,column].map(value_map)

		#Create mapping to be able to convert back
		mapping = dict()
		for column, value_dict in integer_strings_by_column.items():
			for label, integer in value_dict.items():
				mapping[integer] = label
		self.labels_for_cut_dimensions = integer_strings_by_column
		self.integers_for_cut_dimensions = values_by_column
		self.dimension_integer_mapping = {
			'integers': sorted(mapping.keys()),
			'values': [mapping[key] for key in sorted(mapping.keys())] 
		}
		assert self.format_string.format(0) not in self.dimension_integer_mapping['integers']

		return df

	def get_integer_string_mapping(self, dimension):
		if self.config is not None and dimension == self.config.default_dimension_title:
			return {'integer_strings':[],'labels':[]}
		assert dimension in self.labels_for_cut_dimensions, "Dimension " + dimension + " not present"
		mapping_as_dict = self.labels_for_cut_dimensions[dimension]
		sorted_labels = sorted(mapping_as_dict.keys())
		return {'integer_strings':[mapping_as_dict[label] for label in sorted_labels],'labels':sorted_labels}

	def create_row_column_headers(self, df, cuts):
		df['row_heading'] = df.apply( lambda row: self.concat_row_items(row,cuts),axis=1)
		df['column_heading'] = df.apply(lambda x: '%s;%s' % (x['question_code'],x['result_type']),axis=1)
		assert df.row_heading.apply(len).min() > 0, "Blank row header created from cuts " + str(cuts) + "\n" + str(df)
		assert df.column_heading.apply(len).min() > 0, "Blank col header created from cuts " + str(cuts) + "\n" + str(df)
		return df

	def concat_row_items(self,row, columns):
		items_in_order = list()
		for column in columns:
			if column == None:
				items_in_order.append(self.zero_integer_string)
			else:
				assert column in row
				items_in_order.append(str(row[column]))
		return ";".join(items_in_order)

	def flush_aggregation_to_file(self,file_name,df):
		assert {'row_heading','column_heading','aggregation_value'} <= set(df.columns)
		with open(file_name,"a+") as f:
			for r, row in df.iterrows():
				aggregation_value = row['aggregation_value']
				if type(aggregation_value) is not str and np.isnan(aggregation_value):
					aggregation_value = ''
				f.write(str(row['row_heading'])+',' + str(row['column_heading']) + ',' + str(aggregation_value) + '\n')

	def setup_flush_file(self,file_name):
		with open(file_name,"w") as f:
				f.write("row_heading,column_heading,aggregation_value\n")

	def export_cuts_to_files(self,**kwargs):
		assert self.config != None
		self.responses = self.remove_questions_not_used(self.responses)
		self.hist_responses = self.remove_questions_not_used(self.hist_responses)
		for_historical = kwargs.pop('for_historical',False)
		cut_sets = self.config.cuts_to_be_created(for_historical=for_historical)
		self.result_types = self.config.config['result_types']

		responses = self.responses
		demographic_data = self.demographic_data
		if for_historical:
			responses = self.hist_responses
			demographic_data = self.hist_demographic_data

		calculator = NumericOutputCalculator.NumericOutputCalculator(responses=responses,demographic_data=demographic_data,responses_transformed=True)
		for i, cut_set in enumerate(cut_sets):
			print("\rCompleted {0:.0f} % of basic computations. Currently working on {1}".format(i/len(cut_sets)*100,str(cut_set)),end=" ")

			df = pd.DataFrame()
			if 'composite_questions' in self.config.config:
				df = self.compute_aggregation(cut_demographic=cut_set,
												composite_questions=self.config.config['composite_questions'],
												calculator=calculator,
												demographic_data=demographic_data,
												)
			else:
				df = self.compute_aggregation(cut_demographic=cut_set,
												calculator=calculator,
												demographic_data=demographic_data,
												)
			file_name = "cut_" + str(i+1) + ".csv"
			df.to_csv(file_name)
			gc.collect()


	# @profile
	def compute_cuts_from_config(self,**kwargs):
		assert self.config != None
		for_historical = kwargs.pop('for_historical',False)
		file_name = kwargs.pop('flush_file',None)
		if file_name is not None:
			self.setup_flush_file(file_name)
		#Remove responses for questions not used
		self.responses = self.remove_questions_not_used(self.responses)
		self.hist_responses = self.remove_questions_not_used(self.hist_responses)

		# assert type(self.config) == ConfigurationReader.ConfigurationReader
		all_aggregations = list()
		# logging.debug("All cuts are " + str(self.config.cuts_to_be_created()))
		#Set default question as not doing that seems to cause issues
		if 'show_questions' in self.config.config:
			self.default_question = self.config.config['show_questions'][0]
		if 'result_types' in self.config.config:
			self.result_types = self.config.config['result_types']
			self.default_result_type = self.result_types[0]
		cut_sets = self.config.cuts_to_be_created(for_historical=for_historical)
		logging.debug("Cut sets for regular cuts are " + str(cut_sets))
		responses = self.responses
		demographic_data = self.demographic_data
		if for_historical:
			responses = self.hist_responses
			demographic_data = self.hist_demographic_data
		calculator = NumericOutputCalculator.NumericOutputCalculator(responses=responses,demographic_data=demographic_data,responses_transformed=True)
		for i, cut_set in enumerate(cut_sets):
			# logging.debug("Cut set is " + str(cut_set))
			print("\rCompleted {0:.0f} % of basic computations. Currently working on {1}".format(i/len(cut_sets)*100,str(cut_set)),end=" ")
			df = pd.DataFrame()
			if 'composite_questions' in self.config.config:
				df = self.compute_aggregation(cut_demographic=cut_set,
												composite_questions=self.config.config['composite_questions'],
												calculator=calculator,
												demographic_data=demographic_data,
												)
			else:
				df = self.compute_aggregation(cut_demographic=cut_set,
												calculator=calculator,
												demographic_data=demographic_data,
												)

			#Remove samples sizes for questions that don't need them
			assert 'question_code' in df.columns
			# logging.debug("question_code dtype is " + str(df.question_code.dtype))
			# logging.debug("question_codes are " + str(df.question_code))
			questions_to_show_sample_size = df.question_code.unique().tolist()
			if 'show_sample_size_for_questions' in self.config.config:
				questions_to_show_sample_size = self.config.config['show_sample_size_for_questions']
			df = df.ix[(df.result_type != 'sample_size') | df.question_code.isin(questions_to_show_sample_size),:]
			
			df = self.replace_dimensions_with_integers(df)
			df = self.create_row_column_headers(df,cuts=cut_set)
			if file_name is not None:
				self.flush_aggregation_to_file(file_name,pd.DataFrame(df,columns=['row_heading','column_heading','aggregation_value']))
			else:
				all_aggregations.append(pd.DataFrame(df,columns=['row_heading','column_heading','aggregation_value']))
			gc.collect()
		
		# return_table.row_heading = return_table.row_heading.map(self.adjust_zero_padding_of_heading)
		# return_table.column_heading = return_table.column_heading.map(self.adjust_zero_padding_of_heading)
		# assert len(return_table.row_heading.apply(len).unique()) == 1, "Not all row headings have the same length\n" + str(return_table.row_heading.unique())
		# assert len(return_table.column_heading.apply(len).unique()) == 1, "Not all column headings have the same length\n" + str(return_table.column_heading.unique())
		if file_name is None:
			return_table = pd.concat(all_aggregations)
			return return_table.drop_duplicates()

	def remove_questions_not_used(self,df):
		if type(df) != pd.DataFrame or'question_code' not in df.columns:
			return df
		questions_to_show = df.question_code.unique().tolist()
		if 'show_questions' in self.config.config:
			questions_to_show = self.config.config['show_questions']
		df = df.ix[df.question_code.isin(questions_to_show),:]
		return df

	# @profile
	def compute_significance_from_config(self,**kwargs):
		assert self.config != None
		for_historical = kwargs.pop('for_historical',False)
		all_aggregations = list()
		file_name = kwargs.pop('flush_file',None)
		if file_name is not None:
			self.setup_flush_file(file_name)
		#Set default question as not doing that seems to cause issues
		if 'show_questions' in self.config.config:
			self.default_question = self.config.config['show_questions'][0]
		self.default_result_type = 'significance_value'

		no_stat_significance_computation = False
		if 'no_stat_significance_computation' in self.config.config:
			 no_stat_significance_computation = True

		cut_sets = self.config.cuts_to_be_created(for_historical=for_historical)
		logging.debug("Cut sets for significance are " + str(cut_sets))
		responses = self.responses
		demographic_data = self.demographic_data
		if for_historical:
			responses = self.hist_responses
			demographic_data = self.hist_demographic_data
		calculator = NumericOutputCalculator.NumericOutputCalculator(responses=responses,demographic_data=demographic_data,responses_transformed=True)
		for i, cut_set in enumerate(cut_sets):
			df = pd.DataFrame()
			print("\rCompleted {0:.0f} % of significance computations. Currently working on {1}".format(i/len(cut_sets)*100,str(cut_set)),end=" ")
			if 'composite_questions' in self.config.config:
				df = self.compute_significance(cut_demographic=cut_set,
											composite_questions=self.config.config['composite_questions'],
											no_stat_significance_computation=no_stat_significance_computation,
											calculator = calculator,
											demographic_data=demographic_data,
											)
			else:
				df = self.compute_significance(cut_demographic=cut_set, 
												no_stat_significance_computation=no_stat_significance_computation,
												calculator = calculator,
												demographic_data=demographic_data,
												)			
			#Remove samples sizes for questions that don't need them
			assert 'question_code' in df.columns
			# logging.debug("question_code dtype is " + str(df.question_code.dtype))
			# logging.debug("question_codes are " + str(df.question_code))

			#Remove questions that aren't specified
			# questions_to_show = df.question_code.unique().tolist()
			# if 'show_questions' in self.config.config:
			# 	questions_to_show = self.config.config['show_questions']
			# df = df.ix[df.question_code.isin(questions_to_show),:]
			df = self.replace_dimensions_with_integers(df)
			df = self.create_row_column_headers(df,cuts=cut_set)
			if file_name is not None:
				self.flush_aggregation_to_file(file_name,pd.DataFrame(df,columns=['row_heading','column_heading','aggregation_value']))
			else:
				all_aggregations.append(pd.DataFrame(df,columns=['row_heading','column_heading','aggregation_value']))
			gc.collect()
		# return_table.row_heading = return_table.row_heading.map(self.adjust_zero_padding_of_heading)
		# return_table.column_heading = return_table.column_heading.map(self.adjust_zero_padding_of_heading)
		# assert len(return_table.row_heading.apply(len).unique()) == 1, "Not all row headings have the same length\n" + str(return_table.row_heading.unique())
		# assert len(return_table.column_heading.apply(len).unique()) == 1, "Not all column headings have the same length\n" + str(return_table.column_heading.unique())
		if file_name is None:
			return_table = pd.concat(all_aggregations)
			return return_table.drop_duplicates()

	def set_up_excel_spreadsheet(self, workbook):
		orig_wb = load_workbook(filename = 'workbook', use_iterators = True)
		new_wb = Workbook(optimized_write = True)
		for sheet_name in orig_wb.get_sheet_names():
			ws = new_wb.create_sheet()
			ws.title = sheet_name
		return new_wb

	def export_to_excel(self):
		# assert type(self.config) == ConfigurationReader.ConfigurationReader
		assert 'excel_template_file' in self.config.config
		self.ensure_combination_for_every_set_of_demographics = True
		filename = self.config.config['excel_template_file']
		
		compute_historical = self.compute_historical
		#Output display values
		self.compute_cuts_from_config(flush_file='df_dv.csv')
		# output_df = pd.read_csv('df_dv.csv').set_index(['row_heading','column_heading'])
		# # logging.debug("Snapshot of master table " + str(output_df.head()))
		# if not output_df.index.is_unique:
		# 	df = output_df.reset_index()
		# 	duplicate_index_df = df.ix[df.duplicated(cols=['row_heading','column_heading']),:].set_index(['row_heading','column_heading'])
		# 	logging.warning("Duplicate headers found including: " + str(output_df[output_df.index.isin(duplicate_index_df.index)].head()))
		# output_df.reset_index().drop_duplicates(cols=['row_heading','column_heading'],take_last=False).to_csv('df_dv.csv',index=False)
		gc.collect()
		
		#Output significance values
		self.compute_significance_from_config(flush_file='df_sig.csv')
		# output_df = pd.read_csv('df_sig.csv').set_index(['row_heading','column_heading'])
		# logging.debug("Snapshot of master table " + str(output_df.head()))
		# if not output_df.index.is_unique:
		# 	df = output_df.reset_index()
		# 	duplicate_index_df = df.ix[df.duplicated(cols=['row_heading','column_heading']),:].set_index(['row_heading','column_heading'])
		# 	logging.warning("Duplicate headers found including: " + str(output_df[output_df.index.isin(duplicate_index_df.index)].head()))
		# output_df.reset_index().drop_duplicates(cols=['row_heading','column_heading'],take_last=False).to_csv('df_sig.csv',index=False)
		gc.collect()

		if compute_historical:
			print("\nNow working on historical calculations")
			#Output hist display values
			self.compute_cuts_from_config(for_historical=True,flush_file='df_dv_hist.csv')
			# output_df = pd.read_csv('df_dv_hist.csv').set_index(['row_heading','column_heading'])
			# if not output_df.index.is_unique:
			# 	df = output_df.reset_index()
			# 	duplicate_index_df = df.ix[df.duplicated(cols=['row_heading','column_heading']),:].set_index(['row_heading','column_heading'])
			# 	logging.warning("Duplicate headers found including: " + str(output_df[output_df.index.isin(duplicate_index_df.index)].head()))
			# output_df.reset_index().drop_duplicates(cols=['row_heading','column_heading'],take_last=False).to_csv('df_dv_hist.csv',index=False)
			gc.collect()
			
			#Output hist significance values
			self.compute_significance_from_config(for_historical=True,flush_file='df_sig_hist.csv')
			# output_df = pd.read_csv('df_sig_hist.csv').set_index(['row_heading','column_heading'])
			# if not output_df.index.is_unique:
			# 	df = output_df.reset_index()
			# 	duplicate_index_df = df.ix[df.duplicated(cols=['row_heading','column_heading']),:].set_index(['row_heading','column_heading'])
			# 	logging.warning("Duplicate headers found including: " + str(output_df[output_df.index.isin(duplicate_index_df.index)].head()))
			# output_df.reset_index().drop_duplicates(cols=['row_heading','column_heading'],take_last=False).to_csv('df_sig_hist.csv',index=False)
			gc.collect()

		#Set up new workbook
		orig_wb = load_workbook(filename = filename, use_iterators = True)
		new_wb = Workbook(optimized_write = True)
		reserved_sheet_names = ['Lookups','DisplayValues','SignificanceValues','HistDisplayValues','HistSignificanceValues']
		for sheet_name in orig_wb.get_sheet_names():
			if sheet_name not in reserved_sheet_names:
				ws = new_wb.create_sheet()
				ws.title = sheet_name

		#Adjust headings and output
		df_dv = pd.read_csv('df_dv.csv')
		os.remove('df_dv.csv')
		#Check that there are not duplicates
		if not df_dv.set_index(['row_heading','column_heading']).index.is_unique:
			df = df_dv
			duplicate_index_df = df.ix[df.duplicated(cols=['row_heading','column_heading']),:].set_index(['row_heading','column_heading'])
			logging.warning("Duplicate headers found including: " + str(df_dv[df_dv.index.isin(duplicate_index_df.index)].head()))
		df_dv = df_dv.drop_duplicates(cols=['row_heading','column_heading'],take_last=False)
		#Adjust headings
		df_dv.row_heading = df_dv.row_heading.astype(str)
		df_dv.column_heading = df_dv.column_heading.astype(str)
		df_dv.row_heading = df_dv.row_heading.map(self.adjust_zero_padding_of_heading)
		df_dv.column_heading = df_dv.column_heading.map(self.adjust_zero_padding_of_heading)
		#Determine length of heading so we can ensure that it is consistent
		assert len(df_dv.row_heading.apply(len).unique()) == 1, "Not all row headings have the same length\n" + str(df_dv.row_heading.unique())
		assert len(df_dv.column_heading.apply(len).unique()) == 1, "Not all column headings have the same length\n" + str(df_dv.column_heading.unique())
		row_heading_length = df_dv.row_heading.apply(len).get(0)
		column_heading_length = df_dv.column_heading.apply(len).get(0)
		df_dv = df_dv.set_index(['row_heading','column_heading'])
		gc.collect()
		output_series = pd.Series(df_dv['aggregation_value'],index = df_dv.index)
		gc.collect()
		unstacked_values = output_series.unstack()
		gc.collect()
		print("Exporting display values to csv")
		unstacked_values.to_csv('display_values.csv')
		gc.collect()
		print("Copying data from csv to excel workbook")
		self.copy_sheet_to_optimized_workbook('display_values.csv','DisplayValues',new_wb,'disp_value')
		gc.collect()
		os.remove('display_values.csv')

		df_sig = pd.read_csv('df_sig.csv')
		os.remove('df_sig.csv')
		#Check that there are not duplicates
		if not df_sig.set_index(['row_heading','column_heading']).index.is_unique:
			df = df_sig
			duplicate_index_df = df.ix[df.duplicated(cols=['row_heading','column_heading']),:].set_index(['row_heading','column_heading'])
			logging.warning("Duplicate headers found including: " + str(df_sig[df_sig.index.isin(duplicate_index_df.index)].head()))
		df_sig = df_sig.drop_duplicates(cols=['row_heading','column_heading'],take_last=False)
		#Adjust headings
		df_sig.row_heading = df_sig.row_heading.astype(str)
		df_sig.column_heading = df_sig.column_heading.astype(str)
		df_sig.row_heading = df_sig.row_heading.map(self.adjust_zero_padding_of_heading)
		df_sig.column_heading = df_sig.column_heading.map(self.adjust_zero_padding_of_heading)
		assert (df_sig.row_heading.apply(len) == row_heading_length).all(), "Not all row headings have the same length\n" + str(df_sig.row_heading.unique())
		assert (df_sig.column_heading.apply(len) == column_heading_length).all(), "Not all column headings have the same length\n" + str(df_sig.column_heading.unique())
		df_sig = df_sig.set_index(['row_heading','column_heading'])
		gc.collect()
		output_series = pd.Series(df_sig['aggregation_value'],index = df_sig.index)
		gc.collect()
		unstacked_values = output_series.unstack()
		gc.collect()
		print("Exporting significance values to csv")
		unstacked_values.to_csv('significance_values.csv')
		gc.collect()
		print("Copying data from csv to excel workbook")
		self.copy_sheet_to_optimized_workbook('significance_values.csv','SignificanceValues',new_wb,'sig_value')
		gc.collect()
		os.remove('significance_values.csv')

		if compute_historical:
			df_dv_hist = pd.read_csv('df_dv_hist.csv')
			os.remove('df_dv_hist.csv')
			#Check that there are not duplicates
			if not df_dv_hist.set_index(['row_heading','column_heading']).index.is_unique:
				df = df_dv_hist
				duplicate_index_df = df.ix[df.duplicated(cols=['row_heading','column_heading']),:].set_index(['row_heading','column_heading'])
				logging.warning("Duplicate headers found including: " + str(df_dv_hist[df_dv_hist.index.isin(duplicate_index_df.index)].head()))
			df_dv_hist = df_dv_hist.drop_duplicates(cols=['row_heading','column_heading'],take_last=False)
			#Adjust headings
			df_dv_hist.row_heading = df_dv_hist.row_heading.astype(str)
			df_dv_hist.column_heading = df_dv_hist.column_heading.astype(str)
			assert len(df_dv_hist.row_heading.apply(len).unique()) == 1, "Not all row headings have the same length\n" + str(df_dv_hist.row_heading.unique())
			assert len(df_dv_hist.column_heading.apply(len).unique()) == 1, "Not all column headings have the same length\n" + str(df_dv_hist.column_heading.unique())
			row_heading_length = df_dv_hist.row_heading.apply(len).get(0)
			column_heading_length = df_dv_hist.column_heading.apply(len).get(0)
			df_dv_hist = df_dv_hist.set_index(['row_heading','column_heading'])
			gc.collect()
			output_series = pd.Series(df_dv_hist['aggregation_value'],index = df_dv_hist.index)
			gc.collect()
			unstacked_values = output_series.unstack()
			gc.collect()
			print("Exporting historical display values to csv")
			unstacked_values.to_csv('display_values.csv')
			gc.collect()
			print("Copying data from csv to excel workbook")
			self.copy_sheet_to_optimized_workbook('display_values.csv','HistDisplayValues',new_wb,'hist_disp_value')
			gc.collect()
			os.remove('display_values.csv')

			df_sig_hist = pd.read_csv('df_sig_hist.csv')
			os.remove('df_sig_hist.csv')
			#Check that there are not duplicates
			if not df_sig_hist.set_index(['row_heading','column_heading']).index.is_unique:
				df = df_sig_hist
				duplicate_index_df = df.ix[df.duplicated(cols=['row_heading','column_heading']),:].set_index(['row_heading','column_heading'])
				logging.warning("Duplicate headers found including: " + str(df_sig_hist[df_sig_hist.index.isin(duplicate_index_df.index)].head()))
			df_sig_hist = df_sig_hist.drop_duplicates(cols=['row_heading','column_heading'],take_last=False)
			#Adjust headings
			df_sig_hist.row_heading = df_sig_hist.row_heading.astype(str)
			df_sig_hist.column_heading = df_sig_hist.column_heading.astype(str)
			df_sig_hist.row_heading = df_sig_hist.row_heading.map(self.adjust_zero_padding_of_heading)
			df_sig_hist.column_heading = df_sig_hist.column_heading.map(self.adjust_zero_padding_of_heading)
			assert (df_sig_hist.row_heading.apply(len) == row_heading_length).all(), "Not all row headings have the same length\n" + str(df_sig_hist.row_heading.unique())
			assert (df_sig_hist.column_heading.apply(len) == column_heading_length).all(), "Not all column headings have the same length\n" + str(df_sig_hist.column_heading.unique())
			df_sig_hist = df_sig_hist.set_index(['row_heading','column_heading'])
			gc.collect()
			output_series = pd.Series(df_sig_hist['aggregation_value'],index = df_sig_hist.index)
			gc.collect()
			unstacked_values = output_series.unstack()
			gc.collect()
			print("Exporting historical significance values to csv")
			unstacked_values.to_csv('significance_values.csv')
			gc.collect()
			print("Copying data from csv to excel workbook")
			self.copy_sheet_to_optimized_workbook('significance_values.csv','HistSignificanceValues',new_wb,'hist_sig_value')
			gc.collect()
			os.remove('significance_values.csv')

		# wb = load_workbook(filename)

		#Remove existing named ranges
		# for range_name in wb.get_named_ranges():
		# 	wb.remove_named_range(range_name)

		#Add ranges for display_values tab
		# dv_ws = wb.get_sheet_by_name('DisplayValues')
		# range_width = dv_ws.get_highest_column() -1
		# range_height = dv_ws.get_highest_row() - 1
		# wb.create_named_range('disp_value_col_head',dv_ws,self.rc_to_range(row=0,col=1,width=range_width,height=1))
		# wb.create_named_range('disp_value_row_head',dv_ws,self.rc_to_range(row=1,col=0,width=1,height=range_height))
		# wb.create_named_range('disp_value_values',dv_ws,self.rc_to_range(row=1,col=1,width=range_width,height=range_height))

		# #Add ranges for significance_values tab
		# dv_ws = wb.get_sheet_by_name('SignificanceValues')
		# range_width = dv_ws.get_highest_column() -1
		# range_height = dv_ws.get_highest_row() - 1
		# wb.create_named_range('sig_value_col_head',dv_ws,self.rc_to_range(row=0,col=1,width=range_width,height=1))
		# wb.create_named_range('sig_value_row_head',dv_ws,self.rc_to_range(row=1,col=0,width=1,height=range_height))
		# wb.create_named_range('sig_value_values',dv_ws,self.rc_to_range(row=1,col=1,width=range_width,height=range_height))

		# if compute_historical:
		# 	#Add ranges for display_values tab
		# 	dv_ws = wb.get_sheet_by_name('HistDisplayValues')
		# 	range_width = dv_ws.get_highest_column() -1
		# 	range_height = dv_ws.get_highest_row() - 1
		# 	wb.create_named_range('hist_disp_value_col_head',dv_ws,self.rc_to_range(row=0,col=1,width=range_width,height=1))
		# 	wb.create_named_range('hist_disp_value_row_head',dv_ws,self.rc_to_range(row=1,col=0,width=1,height=range_height))
		# 	wb.create_named_range('hist_disp_value_values',dv_ws,self.rc_to_range(row=1,col=1,width=range_width,height=range_height))

		# 	#Add ranges for significance_values tab
		# 	dv_ws = wb.get_sheet_by_name('HistSignificanceValues')
		# 	range_width = dv_ws.get_highest_column() -1
		# 	range_height = dv_ws.get_highest_row() - 1
		# 	wb.create_named_range('hist_sig_value_col_head',dv_ws,self.rc_to_range(row=0,col=1,width=range_width,height=1))
		# 	wb.create_named_range('hist_sig_value_row_head',dv_ws,self.rc_to_range(row=1,col=0,width=1,height=range_height))
		# 	wb.create_named_range('hist_sig_value_values',dv_ws,self.rc_to_range(row=1,col=1,width=range_width,height=range_height))

		# if 'Lookups' in wb.get_sheet_names():
		# 	ws_to_del = wb.get_sheet_by_name('Lookups')
		# 	wb.remove_sheet(ws_to_del)

		print("Adding lookups tab")

		lookup_wb = Workbook()

		ws = lookup_wb.create_sheet()
		ws.title = 'Lookups'
		cuts_menus = self.config.cuts_for_excel_menu(menu=None)
		max_menu_length = max([len(menu) for menu in cuts_menus])
		for menu_i, cut_menu in enumerate(cuts_menus):
			for col_i, item in enumerate(cut_menu):
				ws.cell(row=menu_i +1, column = col_i +1).value = item

		#Added ranges for cut menu
		range_width = ws.get_highest_column() - 1
		range_height = ws.get_highest_row() -1 
		lookup_wb.create_named_range('cuts_config',ws,self.rc_to_range(row=1,col=1,width=range_width + 1,height=range_height + 1))
		
		highest_column = ws.get_highest_column() + 1
		for menu_i, cut_menu in enumerate(self.cut_menu_order(cuts_menus)):
			ws.cell(row=menu_i +1, column = highest_column).value = cut_menu
		lookup_wb.create_named_range('cuts',ws,self.rc_to_range(row=1,col=highest_column,width=1,height=range_height + 1))

		#Add historical cut menu
		menu_start = ws.get_highest_row()+1
		menu = self.config.cuts_for_excel_menu(menu='historical')
		menu_length = len(menu)
		for menu_i, cut_menu in enumerate(self.cut_menu_order(menu)):
			ws.cell(row=menu_i+menu_start, column = highest_column).value = cut_menu
		lookup_wb.create_named_range('cuts_historical',ws,self.rc_to_range(row=menu_start,col=highest_column,width=1,height=menu_length))

		#Add cut menu 2 and 3
		menu_start = ws.get_highest_row()+1
		menu = self.config.cuts_for_excel_menu(menu='cuts_2')
		menu_length = len(menu)
		for menu_i, cut_menu in enumerate(self.cut_menu_order(menu)):
			ws.cell(row=menu_i+menu_start, column = highest_column).value = cut_menu
		lookup_wb.create_named_range('cuts_2',ws,self.rc_to_range(row=menu_start,col=highest_column,width=1,height=menu_length))

		menu_start = ws.get_highest_row()+1
		menu = self.config.cuts_for_excel_menu(menu='cuts_3')
		menu_length = len(menu)
		for menu_i, cut_menu in enumerate(self.cut_menu_order(menu)):
			ws.cell(row=menu_i+menu_start, column = highest_column).value = cut_menu
		lookup_wb.create_named_range('cuts_3',ws,self.rc_to_range(row=menu_start,col=highest_column,width=1,height=menu_length))

		menu_start = ws.get_highest_row()+1
		menu = self.config.cuts_for_excel_menu(menu='cuts_4')
		menu_length = len(menu)
		for menu_i, cut_menu in enumerate(self.cut_menu_order(menu)):
			ws.cell(row=menu_i+menu_start, column = highest_column).value = cut_menu
		lookup_wb.create_named_range('cuts_4',ws,self.rc_to_range(row=menu_start,col=highest_column,width=1,height=menu_length))

		menu_start = ws.get_highest_row()+1
		menu = self.config.cuts_for_excel_menu(menu='cuts_5')
		menu_length = len(menu)
		for menu_i, cut_menu in enumerate(self.cut_menu_order(menu)):
			ws.cell(row=menu_i+menu_start, column = highest_column).value = cut_menu
		lookup_wb.create_named_range('cuts_5',ws,self.rc_to_range(row=menu_start,col=highest_column,width=1,height=menu_length))


		#Add dimension menus
		next_column_to_use = ws.get_highest_column() + 1
		#TODO: Perhaps having a function call directly from config would be better than these structures?
		all_dimensions = {dimension.title: dimension for dimension in self.config.all_dimensions()}
		dimension_titles = [dimension.title for dimension in self.config.all_dimensions()]
		all_together_label_for_title = {dimension.title: dimension.all_together_label for dimension in self.config.all_dimensions()}
		dynamic_parent_dimension = {dimension.title: dimension.dynamic_parent_dimension for dimension in self.config.all_dimensions() if dimension.dimension_type=='dynamic'}
		dimension_titles.append('question_code')
		dimension_titles.append('result_type')

		#Get responses for dynamic dimensions
		responses = self.responses
		demographic_data = self.demographic_data
		responses_for_menu = NumericOutputCalculator.NumericOutputCalculator(responses=responses,demographic_data=demographic_data,responses_transformed=True).responses_with_dimensions		
		responses_for_menu = responses_for_menu.ix[responses_for_menu.response.notnull(),:]
		if compute_historical:
			dimension_titles.append('survey_code')
		for dimension_title in dimension_titles:
			mapping = {'labels':[],'integer_strings':[]}
			if dimension_title in all_dimensions and all_dimensions[dimension_title].is_composite:
				for component_dimension_title in all_dimensions[dimension_title].composite_dimensions:
					new_mapping = self.get_integer_string_mapping(component_dimension_title)
					mapping['labels'] = mapping['labels'] + new_mapping['labels']
					mapping['integer_strings'] = mapping['integer_strings'] + new_mapping['integer_strings']
			else:
				mapping = self.get_integer_string_mapping(dimension_title)

			ws.cell(row=1, column = next_column_to_use).value = dimension_title
			row_offset = 1
			integer_mapping = dict(zip(mapping['labels'],mapping['integer_strings']))
			if dimension_title in dynamic_parent_dimension:
				if dimension_title in all_together_label_for_title and all_together_label_for_title[dimension_title] is not None:
					ws.cell(row=row_offset+1, column = next_column_to_use).value = all_together_label_for_title[dimension_title]
					ws.cell(row=row_offset+1, column = next_column_to_use + 1 ).value = self.zero_integer_string
					row_offset = 2
				assert self.demographic_data is not None, "demographic data not none in dimension " + dimension_title
				parent_dimension = dynamic_parent_dimension[dimension_title]
				assert parent_dimension in self.demographic_data.columns, parent_dimension + " not found in demographic data and is needed for dimension " + dimension_title
				assert dimension_title in self.demographic_data.columns, "Missing dimension " + dimension_title + " in demographic file"
				dimension_mapping = pd.DataFrame(responses_for_menu,columns=[parent_dimension,dimension_title]).drop_duplicates().sort(columns=[parent_dimension,dimension_title])
				i = 0
				for index, row_items in dimension_mapping.iterrows():
					if row_items[dimension_title] in integer_mapping:
						ws.cell(row=i+row_offset+1, column = next_column_to_use).value = row_items[dimension_title]
						ws.cell(row=i+row_offset+1, column = next_column_to_use+1).value = integer_mapping[row_items[dimension_title]]
						ws.cell(row=i+row_offset+1, column = next_column_to_use+2).value = row_items[parent_dimension]
						i += 1
				next_column_to_use += 3
			else:
				if dimension_title in all_together_label_for_title and all_together_label_for_title[dimension_title] is not None:
					ws.cell(row=row_offset+1, column = next_column_to_use).value = all_together_label_for_title[dimension_title]
					ws.cell(row=row_offset+1, column = next_column_to_use + 1 ).value = self.zero_integer_string
					row_offset = 2
				labels_for_menu = mapping['labels']
				if dimension_title in all_dimensions and type(all_dimensions[dimension_title].value_order) is list:
					logging.debug("Writing value order for " + dimension_title)
					labels_for_menu = all_dimensions[dimension_title].value_order
				logging.debug("Labels for menu is " + str(labels_for_menu))
				logging.debug("Integer map for menu is " + str(integer_mapping))
				i = 1
				for label in labels_for_menu:
					if label in integer_mapping:
						ws.cell(row=i+row_offset, column = next_column_to_use).value = label
						ws.cell(row=i+row_offset, column = next_column_to_use+1).value = str(integer_mapping[label])
					i += 1
				next_column_to_use += 2

		#Add ranges for dimension menus
		col_for_default_menu = range_width + 2
		range_width = ws.get_highest_column()
		lookup_wb.create_named_range('default_menu',ws,self.rc_to_range(row=2,col=col_for_default_menu,width=1,height=1000))
		lookup_wb.create_named_range('default_mapping',ws,self.rc_to_range(row=2,col=col_for_default_menu,width=2,height=1000))
		lookup_wb.create_named_range('default_menu_start',ws,self.rc_to_range(row=2,col=col_for_default_menu))
		lookup_wb.create_named_range('cuts_head',ws,self.rc_to_range(row=1,col=col_for_default_menu + 1,
																width=range_width-col_for_default_menu,height=1))

		#Add zero string value
		next_column_to_use = ws.get_highest_column() + 1
		ws.cell(row=1,column=next_column_to_use).value = self.zero_integer_string
		lookup_wb.create_named_range('zero_string',ws,self.rc_to_range(row=1,col=next_column_to_use))

		lookup_wb.save('Lookups.xlsx')
		#Copy lookups to main wb
		lookup_wb = load_workbook(filename = 'Lookups.xlsx', use_iterators = True)
		ws_to_copy = lookup_wb.get_sheet_by_name(name='Lookups')
		lookup_ws = new_wb.create_sheet()
		lookup_ws.title = "Lookups"
		for r, row in enumerate(ws_to_copy.iter_rows()):
			lookup_ws.append([cell.value for cell in row])
			if r % 500 == 0 and r > 0:
				print("\r" + str(r) + " rows for lookups written", end= ' ')
		#Copy ranges names
		for wb_range in lookup_wb.get_named_ranges():
			range_name = wb_range.name
			new_wb.create_named_range(range_name,lookup_ws,lookup_wb.get_named_range(range_name).destinations[0][1])

		# os.remove('Lookups.xlsx')
		gc.collect()
		print("Saving dashboard")
		new_wb.save(filename)
		print("\nExported dashboard")

	def cut_menu_order(self, cuts_menus):
		#Determine default menu order
		if len(cuts_menus) == 0:
			return cuts_menus
		default_order_by_menu = {}
		cut_menus_order = {cut_menu[0]: 0 for cut_menu in cuts_menus}
		if 'cut_menu_order' in self.config.config:
			default_order_by_menu = {menu_title: order for (order, menu_title) in enumerate(self.config.config['cut_menu_order'])}
			logging.debug('Default order by menu is ' + str(default_order_by_menu))
			cut_menus_order = {cut_menu: self.menu_order(cut_menu, default_order_by_menu) for cut_menu in cut_menus_order}
		sorted_cut_menus = sorted(cut_menus_order,key=lambda cut: cut_menus_order[cut])
		logging.debug('Sorted menu is ' + str(sorted_cut_menus))
		return sorted_cut_menus

	def menu_order(self, menu_item, default_menu_order):
		end_of_menu_order = len(default_menu_order)
		if menu_item in default_menu_order:
			return default_menu_order[menu_item]
		else:
			return end_of_menu_order

	def adjust_zero_padding_of_heading(self, input_heading):
		values = input_heading.split(";")
		value_strings = [self.format_string.format(int(value)) for value in values]
		return ";".join(value_strings)

	def rc_to_range(self,row,col,**kwargs):
		height = kwargs.pop('height',None)
		width = kwargs.pop('width',None)
		if height == None:
			return cell.get_column_letter(col) + str(row)
		else:
			assert width != None
			start_cell =  cell.get_column_letter(col) + str(row)
			end_cell = cell.get_column_letter(col + width-1) + str(row+height-1)
			return start_cell + ":" + end_cell

	def copy_sheet_to_optimized_workbook(self,src_data,dest_ws,wb,range_name_prefix):
		assert re.search('.csv$', src_data) is not None
		ws = wb.create_sheet()
		ws.title = dest_ws
		max_col_width = 0
		max_row_width = 0
		with open(src_data) as f:
				reader = csv.reader(f)
				for r,row in enumerate(reader):
					#Set maximums for headers
					max_row_width = r
					if (len(row)-1) > max_col_width:
						max_col_width = (len(row)-1)

					#Add row
					ws.append([self.float_string_conversion(item) for item in row])
					if r % 500 == 0 and r > 0:
						print("\r" + str(r) + " rows written", end= ' ')
		wb.create_named_range(range_name_prefix + '_col_head',ws,self.rc_to_range(row=1,col=2,width=max_col_width,height=1))
		wb.create_named_range(range_name_prefix + '_row_head',ws,self.rc_to_range(row=2,col=1,width=1,height=max_row_width))
		wb.create_named_range(range_name_prefix + '_values',ws,self.rc_to_range(row=2,col=2,width=max_col_width,height=max_row_width))
		return wb

	# @profile
	def copy_sheet_to_workbook(self,src_wb_name,src_ws_name,dest_wb_name):
		dest_wb = load_workbook(dest_wb_name)
		is_csv_file = False
		if re.search('.csv$', src_wb_name) is not None:
			is_csv_file = True

		if src_ws_name in dest_wb.get_sheet_names():
			ws_to_del = dest_wb.get_sheet_by_name(src_ws_name)
			dest_wb.remove_sheet(ws_to_del)

		dest_ws = dest_wb.create_sheet()
		dest_ws.title = src_ws_name

		if is_csv_file:
			with open(src_wb_name) as f:
				reader = csv.reader(f)
				for r,row in enumerate(reader):
					dest_ws.append(row)
					if r % 500 == 0 and r > 0:
						print("\r" + str(r) + " rows written", end= ' ')
						# dest_wb.save(dest_wb_name)
						gc.collect()

		else:
			src_wb = load_workbook(src_wb_name)
			assert src_ws_name in src_wb.get_sheet_names()
			src_ws = src_wb.get_sheet_by_name(src_ws_name)

			for i in range(src_ws.get_highest_column()):
				for j in range(src_ws.get_highest_row()):
					dest_ws.cell(row=j+1,column=i+1).value = src_ws.cell(row=j+1,column=i+1).value
		print("\nSaving file, this may take a few minutes")
		dest_wb.save(dest_wb_name)

	def float_string_conversion(self, string):
		try:
			return float(string)
		except ValueError:
			return string

	def add_pilot_cms(self,pilot_rows):
		assert type(pilot_rows) is list
		if self.config is not None:
			self.config.add_pilot_cuts(pilot_rows)
		pilot_columns = [[row[x] for row in pilot_rows] for x in range(len(pilot_rows[0]))]
		self.demographic_data = self.demographic_data.set_index('respondent_id')
		for col in pilot_columns:
			pilot_name = col[0]
			label_value = col[1]
			ids = col[3:]
			demographic_column_label = pilot_name + "-" + label_value
			self.demographic_data[demographic_column_label] = np.nan
			for respondent_id in ids:
				assert type(respondent_id) is str
				if respondent_id.isdigit():
					try:
						self.demographic_data.ix[int(respondent_id),demographic_column_label] = label_value
					except:
						logging.warning("An error was thrown while setting pilot value for CM " + str(respondent_id) + " for pilot " + pilot_name)
				elif re.match('[\d\.]+',respondent_id) is not None:
					try:
						self.demographic_data.ix[int(float(respondent_id)),demographic_column_label] = label_value
					except:
						logging.warning("An error was thrown while setting pilot value for CM " + str(respondent_id) + " for pilot " + pilot_name)
		self.demographic_data = self.demographic_data.applymap(str).fillna("Missing").reset_index()
